import os
import sqlite3
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font


class SqliteToExcel:
    """
    Converte qualquer banco SQLITE em excel(.xlsx).
    """

    def __init__(self, db_path : str, download_path : str, xlsx_name = None) -> None:
        
        self.db_path = db_path
        self.download_path = download_path
        self.xlsx_name = xlsx_name
        self.con = None
        self._process()
        
    def _process(self):
        with self.connecting_to_db() as self.con:
            cur = self.con.cursor()
            list_tables_name = self.finding_the_names_of_the_tables(cur)
            self.create_excel()
            
            for table_name in list_tables_name:
                fields_name = self.finding_the_names_of_the_fields(cur, table_name)
                data = self.query_all(cur, table_name, fields_name)
                self.list_to_sheet(table_name, fields_name, data)
            
    def connecting_to_db(self) -> sqlite3.connect:
        """
        Cria uma conexão com o banco.
        """
        return sqlite3.connect(self.db_path)
         
    def finding_the_names_of_the_tables(self, cur) -> list:
        """
        Encontra o nome da tabela e retorna o nome dela como uma string.
        """
        cur.execute("SELECT name FROM sqlite_master WHERE type='table';")
        tables_name = cur.fetchall()
        
        return [column[0] for column in tables_name if column[0] != "sqlite_sequence"]
    
    def finding_the_names_of_the_fields(self, cur, table_name) -> list:
        """
        Encontra os nomes dos campos da tabela e retorna uma lista com o nome deles sem o campo PK.
        """
        cur.execute(f"PRAGMA table_info('{table_name}')")
        
        fields_list = cur.fetchall()
        
        # Remove o campo que é primary key
        fields_name = [fields for fields in fields_list if fields[3] != 1]
        
        return [column[1] for column in fields_name]
    
    def query_all(self, cur, table_name, columns) -> list:
        """
        Faz uma consulta de todos os dados da tabela e retorna uma lista de todos eles.
        """
        columns = ', '.join(columns)
        cur.execute(f'SELECT {columns} FROM "{table_name}"')
        return cur.fetchall()
    
    def create_excel(self) -> None:
        """
        Cria um excel
        """
        if self.xlsx_name is None:
            self.download_path = os.path.join(self.download_path, "consultas.xlsx")
        else:
            self.download_path = os.path.join(self.download_path, f"{self.xlsx_name}.xlsx")
        
        wb = Workbook()
        wb.save(self.download_path)
    
    def list_to_sheet(self, table_name, fields_name, data) -> None:
        """
        Passa os dados que estão em listas para dentro de uma Sheet
        """
        load_wb = load_workbook(self.download_path)
        
        # Remove a sheet padrão
        if "Sheet" in load_wb.sheetnames:
            load_wb.remove(load_wb["Sheet"])
        
        load_wb.create_sheet(title=f'{table_name}')
        load_ws = load_wb[table_name]
        
        bold_font = Font(bold=True)
        for i, field in enumerate(fields_name):
            cell = load_ws.cell(row=1, column=i+1, value=field)
            cell.font = bold_font
        
        for row in data:
            load_ws.append(row)
        
        load_wb.save(self.download_path)
        load_wb.close()
        
    def __del__(self) -> None:
        """
        Verifica se a conexão foi fechada.
        """
        if self.con:
            self.con.close()
        
        
if __name__ in "__main__":
    conversor = SqliteToExcel("mia.db", "download_xlsx", "nome_do_excel")
